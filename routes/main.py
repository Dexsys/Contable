import datetime as dt
import json
import os
import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO

from flask import Blueprint, current_app, jsonify, render_template, request, send_file, send_from_directory, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func
from werkzeug.utils import secure_filename

from extensions import db
from models import Account, AuditLog, BankStatement, LedgerEntry, TermDeposit, Voucher, VoucherLine
from routes.security import get_current_user, require_permission

main_bp = Blueprint("main", __name__)

ALLOWED_RECEIPT_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif"}
ALLOWED_STATEMENT_EXTENSIONS = {"pdf", "xlsx"}
MAIN_APPROVERS = ["kazat@colbun.cl", "lcorales@colbun.cl"]


def write_audit_log(action: str, entity: str | None = None, entity_id: int | None = None, detail: str | None = None):
    """Registra un evento en el log de auditoría."""
    user = get_current_user(required=False, auto_create=False)
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()[:45]
    log = AuditLog(
        user_email=user.email if user else None,
        action=action,
        entity=entity,
        entity_id=entity_id,
        detail=detail,
        ip_address=ip or None,
    )
    db.session.add(log)
    # No hacemos commit aquí; el commit lo realiza la transacción principal.


@main_bp.get("/")
def index():
    return render_template("dashboard.html")


@main_bp.get("/health")
def healthcheck():
    return jsonify({"status": "healthy"}), 200


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, dt.date):
        return value

    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha inválida: {value}")


def parse_amount(value, field_name="amount"):
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        raw = str(value).strip().replace("$", "").replace(" ", "")
        if "," in raw and "." in raw:
            if raw.rfind(",") > raw.rfind("."):
                raw = raw.replace(".", "").replace(",", ".")
            else:
                raw = raw.replace(",", "")
        elif "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        return Decimal(raw)
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError(f"Monto inválido para {field_name}") from exc


def allowed_receipt(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_RECEIPT_EXTENSIONS


def allowed_statement(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_STATEMENT_EXTENSIONS


def save_receipt_file(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    filename = secure_filename(file_storage.filename)
    if not allowed_receipt(filename):
        raise ValueError("Formato de imagen no permitido. Usa png, jpg, jpeg, webp o gif")

    ext = filename.rsplit(".", 1)[1].lower()
    unique_name = f"entry_{dt.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.{ext}"

    folder = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(folder, exist_ok=True)
    full_path = os.path.join(folder, unique_name)
    file_storage.save(full_path)
    return unique_name


def save_statement_file(file_storage):
    if not file_storage or not file_storage.filename:
        return None, None

    original_filename = file_storage.filename
    filename = secure_filename(original_filename)
    if not allowed_statement(filename):
        raise ValueError("Formato no permitido. Usa PDF o XLSX")

    ext = filename.rsplit(".", 1)[1].lower()
    unique_name = f"stmt_{dt.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.{ext}"

    folder = current_app.config["UPLOAD_FOLDER"]
    os.makedirs(folder, exist_ok=True)
    full_path = os.path.join(folder, unique_name)
    file_storage.save(full_path)
    return unique_name, ext



def normalize_email(value):
    return (value or "").strip().lower()


def allowed_approvers_for_presenter(presenter_email):
    presenter = normalize_email(presenter_email)
    if presenter in MAIN_APPROVERS:
        return [email for email in MAIN_APPROVERS if email != presenter]
    return list(MAIN_APPROVERS)


def validate_assigned_approver(presenter_email, assigned_approver_email):
    allowed = allowed_approvers_for_presenter(presenter_email)
    assigned = normalize_email(assigned_approver_email)
    if assigned not in allowed:
        raise ValueError("Aprobador no permitido para este presentador")
    return assigned


def next_voucher_number():
    prefix = dt.date.today().strftime("CMP-%Y%m-")
    latest = (
        Voucher.query.filter(Voucher.voucher_number.like(f"{prefix}%"))
        .order_by(Voucher.voucher_number.desc())
        .first()
    )
    if not latest:
        return f"{prefix}0001"
    last = latest.voucher_number.replace(prefix, "")
    try:
        sequence = int(last) + 1
    except ValueError:
        sequence = 1
    return f"{prefix}{sequence:04d}"


def find_account(account_code=None, account_name=None):
    if account_code:
        code = str(account_code).strip().replace(" ", "")
        account = Account.query.filter_by(code=code).one_or_none()
        if account:
            return account
    if account_name:
        return Account.query.filter(func.lower(Account.name) == str(account_name).strip().lower()).one_or_none()
    return None


def build_ledger_entry(payload, movement_type="general", term_deposit_id=None, receipt_image_path=None):
    entry_date = parse_date(payload.get("entry_date") or payload.get("date") or dt.date.today().isoformat())
    bank_effective_date = parse_date(payload.get("bank_effective_date") or entry_date.isoformat())
    description = (payload.get("description") or "").strip()
    if not description:
        raise ValueError("description es requerido")

    kind = (payload.get("kind") or "").strip().lower()
    debit = parse_amount(payload.get("debit"), "debit")
    credit = parse_amount(payload.get("credit"), "credit")
    amount = parse_amount(payload.get("amount"), "amount")

    if amount > 0 and debit == 0 and credit == 0:
        if kind in ("expense", "egreso", "gasto", "term_deposit_open"):
            debit = amount
        elif kind in ("income", "ingreso", "term_deposit_rescue"):
            credit = amount

    if debit == 0 and credit == 0:
        raise ValueError("debit/credit o amount debe ser mayor que cero")

    account = find_account(payload.get("account_code"), payload.get("account_name"))
    raw_code = payload.get("account_code")
    raw_name = payload.get("account_name")

    return LedgerEntry(
        entry_date=entry_date,
        bank_effective_date=bank_effective_date,
        description=description,
        reference=(payload.get("reference") or "").strip() or None,
        debit=debit,
        credit=credit,
        account=account,
        raw_account_code=str(raw_code).strip().replace(" ", "") if raw_code else None,
        raw_account_name=str(raw_name).strip() if raw_name else None,
        note=(payload.get("note") or "").strip() or None,
        movement_type=movement_type,
        term_deposit_id=term_deposit_id,
        receipt_image_path=receipt_image_path,
    )


def parse_voucher_payload():
    if request.is_json:
        payload = request.get_json(silent=True) or {}
        lines = payload.get("lines") or []
    else:
        payload = request.form.to_dict(flat=True)
        lines_raw = request.form.get("lines", "[]")
        try:
            lines = json.loads(lines_raw)
        except json.JSONDecodeError as exc:
            raise ValueError("Formato inválido en lines") from exc

    if not isinstance(lines, list) or len(lines) == 0:
        raise ValueError("Debes enviar al menos una línea de comprobante")

    return payload, lines


def parse_voucher_lines(lines):
    parsed = []
    total_debit = Decimal("0")
    total_credit = Decimal("0")

    for idx, line in enumerate(lines, start=1):
        description = (line.get("description") or "").strip()
        if not description:
            raise ValueError(f"La línea {idx} requiere descripción")

        account_code = (line.get("account_code") or "").strip().replace(" ", "") or None
        account_name = (line.get("account_name") or "").strip() or None
        debit = parse_amount(line.get("debit"), f"linea {idx} debit")
        credit = parse_amount(line.get("credit"), f"linea {idx} credit")

        amount = parse_amount(line.get("amount"), f"linea {idx} amount")
        kind = (line.get("kind") or "").strip().lower()
        if amount > 0 and debit == 0 and credit == 0:
            if kind in ("egreso", "gasto", "expense"):
                debit = amount
            elif kind in ("ingreso", "income"):
                credit = amount

        if debit == 0 and credit == 0:
            raise ValueError(f"La línea {idx} debe tener debit o credit")

        total_debit += debit
        total_credit += credit

        parsed.append(
            {
                "line_number": idx,
                "account_code": account_code,
                "account_name": account_name,
                "description": description,
                "debit": debit,
                "credit": credit,
                "note": (line.get("note") or "").strip() or None,
            }
        )

    if total_debit <= 0 and total_credit <= 0:
        raise ValueError("Comprobante inválido: sin montos")

    return parsed, total_debit, total_credit


def apply_voucher_to_ledger(voucher):
    for line in voucher.lines.order_by(VoucherLine.line_number.asc()).all():
        account = find_account(line.account_code, line.account_name)
        entry = LedgerEntry(
            entry_date=voucher.voucher_date,
            bank_effective_date=voucher.voucher_date,
            description=line.description,
            reference=voucher.voucher_number,
            debit=line.debit,
            credit=line.credit,
            account=account,
            raw_account_code=line.account_code,
            raw_account_name=line.account_name,
            note=line.note or voucher.request_note,
            movement_type="voucher_approved",
            receipt_image_path=voucher.receipt_image_path,
        )
        db.session.add(entry)


def get_period_filters():
    start = parse_date(request.args.get("start_date")) if request.args.get("start_date") else None
    end = parse_date(request.args.get("end_date")) if request.args.get("end_date") else None
    month = request.args.get("month", type=int)
    year = request.args.get("year", type=int)
    return start, end, month, year


def resolve_period_bounds():
    start, end, month, year = get_period_filters()

    if start:
        period_start = start
    elif year and month:
        period_start = dt.date(year, month, 1)
    elif year:
        period_start = dt.date(year, 1, 1)
    else:
        period_start = None

    if end:
        period_end = end
    elif year and month:
        if month == 12:
            period_end = dt.date(year, 12, 31)
        else:
            period_end = dt.date(year, month + 1, 1) - dt.timedelta(days=1)
    elif year:
        period_end = dt.date(year, 12, 31)
    else:
        period_end = None

    return period_start, period_end


def apply_period_filters(query):
    period_start, period_end = resolve_period_bounds()
    filter_date = LedgerEntry.entry_date

    # El periodo siempre se interpreta como un rango cerrado [inicio, fin].
    # Ejemplo: year=2025, month=Todos -> 2025-01-01 a 2025-12-31.
    if period_start:
        query = query.filter(filter_date >= period_start)
    if period_end:
        query = query.filter(filter_date <= period_end)

    return query


def get_entry_account_code_and_name(entry):
    code = entry.account.code if entry.account else entry.raw_account_code
    name = entry.account.name if entry.account else entry.raw_account_name
    code = (code or "sin_cuenta").strip()
    name = (name or "Sin cuenta asignada").strip()
    return code, name


def get_level3_code_and_name(code: str, name: str, names_map: dict[str, str]):
    if code == "sin_cuenta":
        return code, name or "Sin cuenta asignada"

    parts = [part for part in code.split(".") if part]
    if not parts:
        return "sin_cuenta", name or "Sin cuenta asignada"

    level3_code = ".".join(parts[:3])
    resolved_name = names_map.get(level3_code) if level3_code else None
    return level3_code, (resolved_name or name or level3_code or "Sin cuenta asignada")


def aggregate_level3(entries):
    names = account_names_map()
    grouped = {}

    for entry in entries:
        code, name = get_entry_account_code_and_name(entry)
        level3_code, level3_name = get_level3_code_and_name(code, name, names)
        key = level3_code or "sin_cuenta"

        if key not in grouped:
            grouped[key] = {
                "code": key,
                "name": level3_name,
                "debit": Decimal("0"),
                "credit": Decimal("0"),
            }

        grouped[key]["debit"] += Decimal(entry.debit or 0)
        grouped[key]["credit"] += Decimal(entry.credit or 0)

    rows = []
    for item in grouped.values():
        rows.append(
            {
                "code": item["code"],
                "name": item["name"],
                "debit": float(item["debit"]),
                "credit": float(item["credit"]),
                "balance": float(item["credit"] - item["debit"]),
            }
        )

    rows.sort(key=lambda x: x["code"])
    return rows


def build_pending_voucher_summary_items():
    start, end, month, year = get_period_filters()

    query = Voucher.query.filter(Voucher.status == "pending_approval")
    if year:
        query = query.filter(func.extract("year", Voucher.voucher_date) == year)
    if month:
        query = query.filter(func.extract("month", Voucher.voucher_date) == month)
    if start:
        query = query.filter(Voucher.voucher_date >= start)
    if end:
        query = query.filter(Voucher.voucher_date <= end)

    items = []
    for voucher in query.order_by(Voucher.voucher_date.desc(), Voucher.id.desc()).all():
        for line in voucher.lines.order_by(VoucherLine.line_number.asc()).all():
            items.append(
                {
                    "id": f"pending-voucher-{voucher.id}-line-{line.id}",
                    "entry_date": voucher.voucher_date.isoformat(),
                    "bank_effective_date": voucher.voucher_date.isoformat(),
                    "description": line.description,
                    "reference": voucher.voucher_number,
                    "account_code": line.account_code,
                    "account_name": line.account_name,
                    "note": line.note or voucher.request_note,
                    "debit": float(line.debit),
                    "credit": float(line.credit),
                    "movement_type": "voucher_pending",
                    "receipt_image_url": (
                        url_for("main.get_uploaded_file", filename=voucher.receipt_image_path)
                        if voucher.receipt_image_path
                        else None
                    ),
                    "pending": True,
                    "voucher_id": voucher.id,
                    "presenter_email": voucher.presenter_email,
                    "assigned_approver_email": voucher.assigned_approver_email,
                }
            )

    return items


def ensure_tree_node(nodes, code, name):
    if code not in nodes:
        nodes[code] = {
            "code": code,
            "name": name,
            "level": code.count(".") + 1 if code else 0,
            "debit": Decimal("0"),
            "credit": Decimal("0"),
            "balance": Decimal("0"),
            "entry_count": 0,
            "children": [],
            "entries": [],
        }
    return nodes[code]


def account_names_map():
    data = {}
    for acc in Account.query.all():
        data[acc.code] = acc.name
    return data


def summarize_entries(entries, include_entries=False):
    names = account_names_map()
    nodes = {}

    def add_to_node(node, entry):
        node["debit"] += Decimal(entry.debit or 0)
        node["credit"] += Decimal(entry.credit or 0)
        node["balance"] = node["credit"] - node["debit"]
        node["entry_count"] += 1
        if include_entries:
            account_code = entry.account.code if entry.account else entry.raw_account_code
            account_name = entry.account.name if entry.account else entry.raw_account_name
            node["entries"].append(
                {
                    "id": entry.id,
                    "entry_date": entry.entry_date.isoformat(),
                    "bank_effective_date": entry.bank_effective_date.isoformat() if entry.bank_effective_date else None,
                    "description": entry.description,
                    "reference": entry.reference,
                    "account_code": account_code,
                    "account_name": account_name,
                    "note": entry.note,
                    "debit": float(entry.debit),
                    "credit": float(entry.credit),
                    "movement_type": entry.movement_type,
                    "receipt_image_url": (
                        url_for("main.get_uploaded_file", filename=entry.receipt_image_path)
                        if entry.receipt_image_path
                        else None
                    ),
                }
            )

    for entry in entries:
        code = None
        if entry.account and entry.account.code:
            code = entry.account.code
        elif entry.raw_account_code:
            code = entry.raw_account_code
        else:
            code = "sin_cuenta"

        if code == "sin_cuenta":
            node = ensure_tree_node(nodes, code, "Sin cuenta asignada")
            add_to_node(node, entry)
            continue

        parts = code.split(".")
        for idx in range(len(parts)):
            partial = ".".join(parts[: idx + 1])
            fallback_name = names.get(partial) or (entry.raw_account_name if partial == code else f"Nivel {idx + 1}")
            node = ensure_tree_node(nodes, partial, fallback_name)
            add_to_node(node, entry)

    for code, node in nodes.items():
        if code == "sin_cuenta":
            continue
        parent_code = ".".join(code.split(".")[:-1]) if "." in code else None
        if parent_code and parent_code in nodes:
            nodes[parent_code]["children"].append(node)

    roots = [node for code, node in nodes.items() if code == "sin_cuenta" or "." not in code]

    def order_nodes(data):
        data.sort(key=lambda x: x["code"])
        for item in data:
            order_nodes(item["children"])
            item["debit"] = float(item["debit"])
            item["credit"] = float(item["credit"])
            item["balance"] = float(item["balance"])
            if not include_entries:
                item.pop("entries", None)

    order_nodes(roots)
    return roots


@main_bp.post("/api/entries")
@require_permission("create_entries")
def create_entry():
    payload = request.get_json(silent=True) if request.is_json else request.form.to_dict(flat=True)
    payload = payload or {}
    try:
        receipt_image_path = save_receipt_file(request.files.get("receipt_image"))
        entry = build_ledger_entry(payload, receipt_image_path=receipt_image_path)
        db.session.add(entry)
        db.session.flush()
        db.session.commit()
        write_audit_log("entry_created", "ledger_entry", entry.id, entry.description[:200])
        db.session.commit()
        return jsonify(
            {
                "status": "ok",
                "entry_id": entry.id,
                "receipt_image_url": (
                    url_for("main.get_uploaded_file", filename=entry.receipt_image_path)
                    if entry.receipt_image_path
                    else None
                ),
            }
        ), 201
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.patch("/api/entries/<int:entry_id>")
@require_permission("create_entries")
def update_entry(entry_id):
    payload = request.get_json(silent=True) or request.form.to_dict(flat=True)
    payload = payload or {}

    entry = LedgerEntry.query.filter_by(id=entry_id).one_or_none()
    if not entry:
        return jsonify({"status": "error", "message": "Movimiento no encontrado"}), 404

    try:
        entry_date = parse_date(payload.get("entry_date") or payload.get("date") or entry.entry_date)
        bank_effective_date = parse_date(payload.get("bank_effective_date") or entry.bank_effective_date or entry_date)
        description = (payload.get("description") or entry.description or "").strip()
        if not description:
            raise ValueError("description es requerido")

        debit = parse_amount(payload.get("debit", entry.debit), "debit")
        credit = parse_amount(payload.get("credit", entry.credit), "credit")
        if debit == 0 and credit == 0:
            raise ValueError("debit o credit debe ser mayor que cero")

        account_code = payload.get("account_code")
        account_name = payload.get("account_name")
        if account_code is None:
            account_code = entry.account.code if entry.account else entry.raw_account_code
        if account_name is None:
            account_name = entry.account.name if entry.account else entry.raw_account_name

        account_code = str(account_code).strip().replace(" ", "") if account_code else None
        account_name = str(account_name).strip() if account_name else None
        account = find_account(account_code, account_name)

        entry.entry_date = entry_date
        entry.bank_effective_date = bank_effective_date
        entry.description = description
        entry.reference = (payload.get("reference") or entry.reference or "").strip() or None
        entry.debit = debit
        entry.credit = credit
        entry.account = account
        entry.raw_account_code = account_code
        entry.raw_account_name = account_name
        entry.note = (payload.get("note") or entry.note or "").strip() or None

        db.session.commit()
        write_audit_log("entry_updated", "ledger_entry", entry.id, entry.description[:200])
        db.session.commit()
        return jsonify({"status": "ok", "entry_id": entry.id}), 200
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.delete("/api/entries/<int:entry_id>")
@require_permission("create_entries")
def delete_entry(entry_id):
    entry = LedgerEntry.query.filter_by(id=entry_id).one_or_none()
    if not entry:
        return jsonify({"status": "error", "message": "Movimiento no encontrado"}), 404

    write_audit_log("entry_deleted", "ledger_entry", entry.id, entry.description[:200])
    db.session.delete(entry)
    db.session.commit()
    return jsonify({"status": "ok", "entry_id": entry_id}), 200


@main_bp.post("/api/term-deposits/open")
@require_permission("manage_term_deposits")
def open_term_deposit():
    payload = request.get_json(silent=True) or {}
    try:
        code = (payload.get("code") or "").strip()
        if not code:
            raise ValueError("code es requerido")

        opened_at = parse_date(payload.get("opened_at") or dt.date.today().isoformat())
        maturity_at = parse_date(payload.get("maturity_at")) if payload.get("maturity_at") else None
        principal = parse_amount(payload.get("principal_amount"), "principal_amount")
        if principal <= 0:
            raise ValueError("principal_amount debe ser mayor que cero")

        deposit = TermDeposit(
            code=code,
            opened_at=opened_at,
            maturity_at=maturity_at,
            principal_amount=principal,
            status="open",
            institution=(payload.get("institution") or "").strip() or None,
            note=(payload.get("note") or "").strip() or None,
        )
        db.session.add(deposit)
        db.session.flush()

        entry_payload = {
            "entry_date": opened_at.isoformat(),
            "bank_effective_date": opened_at.isoformat(),
            "description": payload.get("description") or f"Apertura deposito a plazo {code}",
            "amount": str(principal),
            "kind": "term_deposit_open",
            "account_code": payload.get("account_code") or "3.01.01",
            "account_name": payload.get("account_name") or "Depositos a Plazo",
            "reference": payload.get("reference"),
            "note": payload.get("note"),
        }
        entry = build_ledger_entry(entry_payload, movement_type="term_deposit_open", term_deposit_id=deposit.id)
        db.session.add(entry)
        db.session.commit()
        write_audit_log("term_deposit_opened", "term_deposit", deposit.id, f"Codigo: {deposit.code}")
        db.session.commit()

        return jsonify({"status": "ok", "term_deposit_id": deposit.id, "entry_id": entry.id}), 201
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.post("/api/term-deposits/<string:code>/rescue")
@require_permission("manage_term_deposits")
def rescue_term_deposit(code):
    payload = request.get_json(silent=True) or {}
    deposit = TermDeposit.query.filter_by(code=code).one_or_none()
    if not deposit:
        return jsonify({"status": "error", "message": "Deposito no encontrado"}), 404

    try:
        rescue_date = parse_date(payload.get("rescued_at") or dt.date.today().isoformat())
        rescue_amount = parse_amount(payload.get("rescue_amount"), "rescue_amount")
        if rescue_amount <= 0:
            raise ValueError("rescue_amount debe ser mayor que cero")

        deposit.rescued_at = rescue_date
        deposit.rescue_amount = rescue_amount
        deposit.status = "rescued"

        entry_payload = {
            "entry_date": rescue_date.isoformat(),
            "bank_effective_date": rescue_date.isoformat(),
            "description": payload.get("description") or f"Rescate deposito a plazo {deposit.code}",
            "amount": str(rescue_amount),
            "kind": "term_deposit_rescue",
            "account_code": payload.get("account_code") or "3.01.03",
            "account_name": payload.get("account_name") or "Rescate Deposito a Plazo",
            "reference": payload.get("reference"),
            "note": payload.get("note"),
        }
        entry = build_ledger_entry(entry_payload, movement_type="term_deposit_rescue", term_deposit_id=deposit.id)
        db.session.add(entry)
        db.session.commit()
        write_audit_log("term_deposit_rescued", "term_deposit", deposit.id, f"Codigo: {deposit.code}")
        db.session.commit()

        return jsonify({"status": "ok", "term_deposit_id": deposit.id, "entry_id": entry.id}), 200
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.get("/api/term-deposits")
@require_permission("view_reports")
def list_term_deposits():
    status = request.args.get("status")
    query = TermDeposit.query
    if status:
        query = query.filter_by(status=status)
    rows = query.order_by(TermDeposit.opened_at.desc()).all()

    return jsonify(
        {
            "items": [
                {
                    "id": row.id,
                    "code": row.code,
                    "opened_at": row.opened_at.isoformat(),
                    "maturity_at": row.maturity_at.isoformat() if row.maturity_at else None,
                    "rescued_at": row.rescued_at.isoformat() if row.rescued_at else None,
                    "principal_amount": float(row.principal_amount),
                    "rescue_amount": float(row.rescue_amount) if row.rescue_amount is not None else None,
                    "status": row.status,
                    "institution": row.institution,
                    "note": row.note,
                }
                for row in rows
            ]
        }
    )


@main_bp.get("/api/reports/bank-summary")
@require_permission("view_reports")
def bank_summary_report():
    include_entries = request.args.get("include_entries", "0") in ("1", "true", "yes")
    period_start, period_end = resolve_period_bounds()
    filter_date = LedgerEntry.entry_date

    query = LedgerEntry.query.order_by(LedgerEntry.entry_date.asc(), LedgerEntry.id.asc())
    query = apply_period_filters(query)
    entries = query.all()

    total_debit = sum((Decimal(e.debit or 0) for e in entries), Decimal("0"))
    total_credit = sum((Decimal(e.credit or 0) for e in entries), Decimal("0"))
    period_balance = total_credit - total_debit

    previous_balance = Decimal("0")
    if period_start:
        previous_query = LedgerEntry.query.filter(filter_date < period_start)
        previous_entries = previous_query.all()
        previous_debit = sum((Decimal(e.debit or 0) for e in previous_entries), Decimal("0"))
        previous_credit = sum((Decimal(e.credit or 0) for e in previous_entries), Decimal("0"))
        previous_balance = previous_credit - previous_debit

    accumulated_final_balance = previous_balance + period_balance
    pending_items = build_pending_voucher_summary_items() if include_entries else []

    return jsonify(
        {
            "filters": {
                "year": request.args.get("year"),
                "month": request.args.get("month"),
                "start_date": request.args.get("start_date"),
                "end_date": request.args.get("end_date"),
                "period_start": period_start.isoformat() if period_start else None,
                "period_end": period_end.isoformat() if period_end else None,
            },
            "totals": {
                "debit": float(total_debit),
                "credit": float(total_credit),
                "balance": float(period_balance),
                "previous_balance": float(previous_balance),
                "period_balance": float(period_balance),
                "accumulated_final_balance": float(accumulated_final_balance),
                "entries": len(entries),
            },
            "tree": summarize_entries(entries, include_entries=include_entries),
            "pending_items": pending_items,
        }
    )


@main_bp.get("/api/reports/available-periods")
@require_permission("view_reports")
def available_periods_report():
    rows = LedgerEntry.query.order_by(LedgerEntry.entry_date.asc(), LedgerEntry.id.asc()).all()

    years = set()
    months_by_year = {}
    latest_date = None

    for row in rows:
        filter_date = row.entry_date
        if not filter_date:
            continue

        year = filter_date.year
        month = filter_date.month
        years.add(year)
        months_by_year.setdefault(str(year), set()).add(month)
        if latest_date is None or filter_date > latest_date:
            latest_date = filter_date

    ordered_years = sorted(years)
    return jsonify(
        {
            "years": ordered_years,
            "months_by_year": {
                year: sorted(list(months))
                for year, months in months_by_year.items()
            },
            "latest": {
                "year": latest_date.year if latest_date else None,
                "month": latest_date.month if latest_date else None,
            },
        }
    )


@main_bp.get("/api/reports/insights")
@require_permission("view_reports")
def report_insights():
    period_start, period_end = resolve_period_bounds()

    period_query = apply_period_filters(LedgerEntry.query)
    period_entries = period_query.order_by(LedgerEntry.entry_date.desc(), LedgerEntry.id.desc()).all()

    accumulated_query = LedgerEntry.query
    if period_end:
        accumulated_query = accumulated_query.filter(LedgerEntry.entry_date <= period_end)
    accumulated_entries = accumulated_query.order_by(LedgerEntry.entry_date.asc(), LedgerEntry.id.asc()).all()

    expense_items = []
    income_items = []

    for entry in period_entries:
        code, name = get_entry_account_code_and_name(entry)
        if Decimal(entry.debit or 0) > 0:
            expense_items.append(
                {
                    "id": entry.id,
                    "entry_date": entry.entry_date.isoformat(),
                    "description": entry.description,
                    "account_code": code,
                    "account_name": name,
                    "amount": float(entry.debit),
                }
            )
        if Decimal(entry.credit or 0) > 0:
            income_items.append(
                {
                    "id": entry.id,
                    "entry_date": entry.entry_date.isoformat(),
                    "description": entry.description,
                    "account_code": code,
                    "account_name": name,
                    "amount": float(entry.credit),
                }
            )

    top_expenses = sorted(expense_items, key=lambda x: x["amount"], reverse=True)[:5]
    top_incomes = sorted(income_items, key=lambda x: x["amount"], reverse=True)[:5]

    investments_by_type_raw = {}
    for entry in accumulated_entries:
        movement_type = (entry.movement_type or "general").strip() or "general"
        if movement_type not in ("term_deposit_open", "term_deposit_rescue"):
            continue

        label = "Aperturas DP" if movement_type == "term_deposit_open" else "Rescates DP"
        value = Decimal(entry.debit or 0) if movement_type == "term_deposit_open" else Decimal(entry.credit or 0)
        if value <= 0:
            continue
        investments_by_type_raw[label] = investments_by_type_raw.get(label, Decimal("0")) + value

    total_open = investments_by_type_raw.get("Aperturas DP", Decimal("0"))
    total_rescue = investments_by_type_raw.get("Rescates DP", Decimal("0"))
    net_invested = total_open - total_rescue

    investments_by_type = [
        {
            "type": label,
            "amount": float(amount),
        }
        for label, amount in sorted(investments_by_type_raw.items(), key=lambda x: x[0])
    ]
    investments_by_type.append({"type": "Neto invertido", "amount": float(net_invested)})

    level3_period = aggregate_level3(period_entries)
    level3_accumulated = aggregate_level3(accumulated_entries)
    period_map = {row["code"]: row for row in level3_period}
    accumulated_map = {row["code"]: row for row in level3_accumulated}

    merged_level3 = []
    all_codes = sorted(set(period_map.keys()) | set(accumulated_map.keys()))
    for code in all_codes:
        period_row = period_map.get(code)
        acc_row = accumulated_map.get(code)
        merged_level3.append(
            {
                "code": code,
                "name": (period_row or acc_row or {}).get("name", code),
                "period": {
                    "debit": (period_row or {}).get("debit", 0.0),
                    "credit": (period_row or {}).get("credit", 0.0),
                    "balance": (period_row or {}).get("balance", 0.0),
                },
                "accumulated": {
                    "debit": (acc_row or {}).get("debit", 0.0),
                    "credit": (acc_row or {}).get("credit", 0.0),
                    "balance": (acc_row or {}).get("balance", 0.0),
                },
            }
        )

    return jsonify(
        {
            "filters": {
                "year": request.args.get("year"),
                "month": request.args.get("month"),
                "start_date": request.args.get("start_date"),
                "end_date": request.args.get("end_date"),
                "period_start": period_start.isoformat() if period_start else None,
                "period_end": period_end.isoformat() if period_end else None,
            },
            "top_expenses": top_expenses,
            "top_incomes": top_incomes,
            "investments_by_type": investments_by_type,
            "level3_summary": merged_level3,
        }
    )


@main_bp.get("/api/accounts")
@require_permission("view_reports")
def list_accounts():
    accounts = Account.query.order_by(Account.code.asc()).all()
    return jsonify(
        {
            "items": [
                {
                    "id": acc.id,
                    "code": acc.code,
                    "name": acc.name,
                    "level": acc.level,
                    "category": acc.category,
                }
                for acc in accounts
            ]
        }
    )


@main_bp.get("/api/entries/recent")
@require_permission("view_reports")
def recent_entries():
    limit = request.args.get("limit", type=int, default=25)
    limit = min(max(limit, 1), 200)

    query = LedgerEntry.query
    query = apply_period_filters(query)
    rows = (
        query.order_by(LedgerEntry.entry_date.desc(), LedgerEntry.id.desc())
        .limit(limit)
        .all()
    )

    return jsonify(
        {
            "items": [
                {
                    "id": row.id,
                    "entry_date": row.entry_date.isoformat(),
                    "description": row.description,
                    "reference": row.reference,
                    "debit": float(row.debit),
                    "credit": float(row.credit),
                    "movement_type": row.movement_type,
                    "account_code": row.account.code if row.account else row.raw_account_code,
                    "account_name": row.account.name if row.account else row.raw_account_name,
                    "receipt_image_url": (
                        url_for("main.get_uploaded_file", filename=row.receipt_image_path)
                        if row.receipt_image_path
                        else None
                    ),
                }
                for row in rows
            ]
        }
    )


@main_bp.get("/uploads/<path:filename>")
@require_permission("view_reports")
def get_uploaded_file(filename):
    safe_name = secure_filename(filename)
    return send_from_directory(current_app.config["UPLOAD_FOLDER"], safe_name)


@main_bp.post("/api/vouchers")
@require_permission("create_vouchers")
def create_voucher():
    try:
        current_user = get_current_user(required=True, auto_create=True)
        payload, lines = parse_voucher_payload()

        presenter_email = normalize_email(payload.get("presenter_email"))
        if not presenter_email and current_user:
            presenter_email = current_user.normalized_email()
        if not presenter_email:
            raise ValueError("presenter_email es requerido")

        # Roles no administrativos solo pueden registrar comprobantes propios.
        if not current_user.can("approve_vouchers") and presenter_email != current_user.normalized_email():
            raise ValueError("No puedes registrar comprobantes para otro usuario")

        requested_approver = normalize_email(payload.get("assigned_approver_email"))
        if not requested_approver:
            requested_approver = allowed_approvers_for_presenter(presenter_email)[0]

        assigned_approver = validate_assigned_approver(presenter_email, requested_approver)
        parsed_lines, total_debit, total_credit = parse_voucher_lines(lines)

        voucher_date = parse_date(payload.get("voucher_date") or payload.get("date") or dt.date.today().isoformat())
        voucher_number = payload.get("voucher_number") or next_voucher_number()
        voucher_number = str(voucher_number).strip().upper()

        image_path = save_receipt_file(request.files.get("receipt_image"))

        voucher = Voucher(
            voucher_number=voucher_number,
            voucher_date=voucher_date,
            presenter_name=(payload.get("presenter_name") or "").strip() or None,
            presenter_email=presenter_email,
            assigned_approver_email=assigned_approver,
            status="pending_approval",
            description=(payload.get("description") or "").strip() or None,
            request_note=(payload.get("request_note") or "").strip() or None,
            receipt_image_path=image_path,
        )
        db.session.add(voucher)
        db.session.flush()

        for line in parsed_lines:
            db.session.add(
                VoucherLine(
                    voucher_id=voucher.id,
                    line_number=line["line_number"],
                    account_code=line["account_code"],
                    account_name=line["account_name"],
                    description=line["description"],
                    debit=line["debit"],
                    credit=line["credit"],
                    note=line["note"],
                )
            )

        db.session.commit()
        write_audit_log("voucher_created", "voucher", voucher.id, f"Nro: {voucher.voucher_number} | Presentador: {voucher.presenter_email}")
        db.session.commit()
        return (
            jsonify(
                {
                    "status": "ok",
                    "voucher_id": voucher.id,
                    "voucher_number": voucher.voucher_number,
                    "assigned_approver_email": voucher.assigned_approver_email,
                    "totals": {"debit": float(total_debit), "credit": float(total_credit)},
                }
            ),
            201,
        )
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.get("/api/vouchers")
@require_permission("view_reports")
def list_vouchers():
    current_user = get_current_user(required=True, auto_create=True)
    status = (request.args.get("status") or "").strip()
    presenter_email = normalize_email(request.args.get("presenter_email"))

    query = Voucher.query
    if status:
        query = query.filter(Voucher.status == status)
    if current_user.can("approve_vouchers"):
        if presenter_email:
            query = query.filter(Voucher.presenter_email == presenter_email)
    else:
        query = query.filter(Voucher.presenter_email == current_user.normalized_email())

    rows = query.order_by(Voucher.created_at.desc()).limit(300).all()
    return jsonify(
        {
            "items": [
                {
                    "id": row.id,
                    "voucher_number": row.voucher_number,
                    "voucher_date": row.voucher_date.isoformat(),
                    "presenter_name": row.presenter_name,
                    "presenter_email": row.presenter_email,
                    "assigned_approver_email": row.assigned_approver_email,
                    "approved_by_email": row.approved_by_email,
                    "rejected_by_email": row.rejected_by_email,
                    "rejection_reason": row.rejection_reason,
                    "status": row.status,
                    "description": row.description,
                    "request_note": row.request_note,
                    "line_count": row.lines.count(),
                    "receipt_image_url": (
                        url_for("main.get_uploaded_file", filename=row.receipt_image_path)
                        if row.receipt_image_path
                        else None
                    ),
                }
                for row in rows
            ]
        }
    )


@main_bp.post("/api/vouchers/<int:voucher_id>/approve")
@require_permission("approve_vouchers")
def approve_voucher(voucher_id):
    payload = request.get_json(silent=True) or request.form.to_dict(flat=True)
    approver_email = normalize_email(payload.get("approver_email"))
    if not approver_email:
        return jsonify({"status": "error", "message": "approver_email es requerido"}), 400

    voucher = Voucher.query.filter_by(id=voucher_id).one_or_none()
    if not voucher:
        return jsonify({"status": "error", "message": "Comprobante no encontrado"}), 404

    if voucher.status != "pending_approval":
        return jsonify({"status": "error", "message": "Comprobante ya procesado"}), 400

    try:
        validate_assigned_approver(voucher.presenter_email, approver_email)

        voucher.status = "approved"
        voucher.approved_by_email = approver_email
        voucher.approved_at = dt.datetime.utcnow()

        apply_voucher_to_ledger(voucher)
        db.session.commit()
        write_audit_log("voucher_approved", "voucher", voucher.id, f"Nro: {voucher.voucher_number} | Aprobador: {approver_email}")
        db.session.commit()

        return jsonify({"status": "ok", "voucher_id": voucher.id, "voucher_number": voucher.voucher_number})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.post("/api/vouchers/<int:voucher_id>/reject")
@require_permission("approve_vouchers")
def reject_voucher(voucher_id):
    payload = request.get_json(silent=True) or request.form.to_dict(flat=True)
    rejector_email = normalize_email(payload.get("rejector_email"))
    rejection_reason = (payload.get("rejection_reason") or "").strip()

    if not rejector_email:
        return jsonify({"status": "error", "message": "rejector_email es requerido"}), 400
    if not rejection_reason:
        return jsonify({"status": "error", "message": "rejection_reason es requerido"}), 400

    voucher = Voucher.query.filter_by(id=voucher_id).one_or_none()
    if not voucher:
        return jsonify({"status": "error", "message": "Comprobante no encontrado"}), 404

    if voucher.status != "pending_approval":
        return jsonify({"status": "error", "message": "Comprobante ya procesado"}), 400

    try:
        validate_assigned_approver(voucher.presenter_email, rejector_email)

        voucher.status = "rejected"
        voucher.rejected_by_email = rejector_email
        voucher.rejected_at = dt.datetime.utcnow()
        voucher.rejection_reason = rejection_reason

        db.session.commit()
        write_audit_log(
            "voucher_rejected",
            "voucher",
            voucher.id,
            f"Nro: {voucher.voucher_number} | Rechazador: {rejector_email} | Motivo: {rejection_reason[:200]}",
        )
        db.session.commit()

        return jsonify({"status": "ok", "voucher_id": voucher.id, "voucher_number": voucher.voucher_number})
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400
    except Exception as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Error interno: {str(exc)}"}), 500


@main_bp.post("/api/bank-statements")
@require_permission("manage_term_deposits")
def upload_bank_statement():
    try:
        year = request.form.get("year", type=int)
        month = request.form.get("month", type=int)
        description = (request.form.get("description") or "").strip()

        if not year or not month or month < 1 or month > 12:
            return jsonify({"status": "error", "message": "año y mes (1-12) son requeridos"}), 400

        current_user = get_current_user(required=True, auto_create=False)
        if not current_user:
            return jsonify({"status": "error", "message": "No autenticado"}), 401

        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"status": "error", "message": "Archivo requerido"}), 400

        filename, ext = save_statement_file(file)

        existing = BankStatement.query.filter_by(year=year, month=month).one_or_none()
        if existing:
            db.session.delete(existing)

        statement = BankStatement(
            year=year,
            month=month,
            filename=filename,
            original_filename=file.filename,
            file_type=ext,
            file_size_bytes=len(file.read()),
            uploaded_by_email=current_user.email,
            description=description,
        )
        file.seek(0)
        statement.file_size_bytes = len(file.read())
        db.session.add(statement)
        write_audit_log("bank_statement_uploaded", "bank_statement", None, f"{year}-{month:02d}")
        db.session.commit()

        return jsonify(
            {
                "status": "ok",
                "message": "Cartola subida",
                "statement": statement.to_dict(),
            }
        ), 201
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(exc)}), 400


@main_bp.get("/api/bank-statements")
@require_permission("view_reports")
def list_bank_statements():
    statements = BankStatement.query.order_by(BankStatement.uploaded_at.desc()).all()
    return jsonify({"items": [s.to_dict() for s in statements]})


@main_bp.delete("/api/bank-statements/<int:statement_id>")
@require_permission("manage_term_deposits")
def delete_bank_statement(statement_id):
    statement = BankStatement.query.filter_by(id=statement_id).one_or_none()
    if not statement:
        return jsonify({"status": "error", "message": "Cartola no encontrada"}), 404

    filename = statement.filename
    folder = current_app.config["UPLOAD_FOLDER"]
    filepath = os.path.join(folder, filename)
    if os.path.exists(filepath):
        os.remove(filepath)

    write_audit_log("bank_statement_deleted", "bank_statement", statement_id, f"{statement.year}-{statement.month:02d}")
    db.session.delete(statement)
    db.session.commit()

    return jsonify({"status": "ok", "message": "Cartola eliminada"}), 200


@main_bp.get("/api/bank-statements/<int:statement_id>/download")
@require_permission("view_reports")
def download_bank_statement(statement_id):
    statement = BankStatement.query.filter_by(id=statement_id).one_or_none()
    if not statement:
        return jsonify({"status": "error", "message": "Cartola no encontrada"}), 404

    folder = current_app.config["UPLOAD_FOLDER"]
    safe_name = secure_filename(statement.filename)
    try:
        return send_from_directory(folder, safe_name, as_attachment=True, download_name=statement.original_filename)
    except FileNotFoundError:
        return jsonify({"status": "error", "message": "Archivo no encontrado en servidor"}), 404


@main_bp.get("/api/reports/export")
@require_permission("view_reports")
def export_report():
    format_type = request.args.get("format", "xlsx").lower()
    if format_type not in ("xlsx",):
        return jsonify({"status": "error", "message": "Formato no soportado. Usa xlsx"}), 400

    period_start, period_end = resolve_period_bounds()

    period_query = apply_period_filters(LedgerEntry.query)
    period_entries = period_query.order_by(LedgerEntry.entry_date.asc(), LedgerEntry.id.asc()).all()

    wb = Workbook()
    wb.remove(wb.active)

    ws_top_expenses = wb.create_sheet("Top 5 Egresos")
    ws_top_incomes = wb.create_sheet("Top 5 Ingresos")
    ws_investments = wb.create_sheet("Inversiones")
    ws_level3 = wb.create_sheet("Resumen Nivel 3")

    header_fill = PatternFill(start_color="1F7A8C", end_color="1F7A8C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_format = '#,##0.00'

    expense_items = []
    income_items = []
    for entry in period_entries:
        code, name = get_entry_account_code_and_name(entry)
        if Decimal(entry.debit or 0) > 0:
            expense_items.append(
                {
                    "description": entry.description,
                    "account": f"{code} - {name}",
                    "amount": float(entry.debit),
                    "date": entry.entry_date.isoformat(),
                }
            )
        if Decimal(entry.credit or 0) > 0:
            income_items.append(
                {
                    "description": entry.description,
                    "account": f"{code} - {name}",
                    "amount": float(entry.credit),
                    "date": entry.entry_date.isoformat(),
                }
            )

    top_expenses = sorted(expense_items, key=lambda x: x["amount"], reverse=True)[:5]
    top_incomes = sorted(income_items, key=lambda x: x["amount"], reverse=True)[:5]

    ws_top_expenses.append(["Fecha", "Descripción", "Cuenta", "Monto"])
    for item in top_expenses:
        ws_top_expenses.append([item["date"], item["description"], item["account"], item["amount"]])

    ws_top_incomes.append(["Fecha", "Descripción", "Cuenta", "Monto"])
    for item in top_incomes:
        ws_top_incomes.append([item["date"], item["description"], item["account"], item["amount"]])

    accumulated_query = LedgerEntry.query
    if period_end:
        accumulated_query = accumulated_query.filter(LedgerEntry.entry_date <= period_end)
    accumulated_entries = accumulated_query.all()

    investments_by_type_raw = {}
    for entry in accumulated_entries:
        movement_type = (entry.movement_type or "general").strip() or "general"
        if movement_type not in ("term_deposit_open", "term_deposit_rescue"):
            continue
        label = "Aperturas DP" if movement_type == "term_deposit_open" else "Rescates DP"
        value = Decimal(entry.debit or 0) if movement_type == "term_deposit_open" else Decimal(entry.credit or 0)
        if value <= 0:
            continue
        investments_by_type_raw[label] = investments_by_type_raw.get(label, Decimal("0")) + value

    total_open = investments_by_type_raw.get("Aperturas DP", Decimal("0"))
    total_rescue = investments_by_type_raw.get("Rescates DP", Decimal("0"))
    net_invested = total_open - total_rescue

    ws_investments.append(["Tipo", "Monto"])
    for label, amount in sorted(investments_by_type_raw.items()):
        ws_investments.append([label, float(amount)])
    ws_investments.append(["Neto invertido", float(net_invested)])

    level3_period = aggregate_level3(period_entries)
    level3_accumulated = aggregate_level3(accumulated_entries)
    period_map = {row["code"]: row for row in level3_period}
    accumulated_map = {row["code"]: row for row in level3_accumulated}

    ws_level3.append(
        [
            "Código",
            "Cuenta",
            "Ingreso Período",
            "Egreso Período",
            "Saldo Período",
            "Ingreso Acumulado",
            "Egreso Acumulado",
            "Saldo Acumulado",
        ]
    )

    all_codes = sorted(set(period_map.keys()) | set(accumulated_map.keys()))
    for code in all_codes:
        period_row = period_map.get(code)
        acc_row = accumulated_map.get(code)
        ws_level3.append(
            [
                code,
                (period_row or acc_row or {}).get("name", code),
                (period_row or {}).get("credit", 0.0),
                (period_row or {}).get("debit", 0.0),
                (period_row or {}).get("balance", 0.0),
                (acc_row or {}).get("credit", 0.0),
                (acc_row or {}).get("debit", 0.0),
                (acc_row or {}).get("balance", 0.0),
            ]
        )

    for ws in wb.sheetnames:
        sheet = wb[ws]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                if isinstance(cell.value, float):
                    cell.number_format = money_format
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws_name in wb.sheetnames:
        sheet = wb[ws_name]
        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 18
        for col in ["E", "F", "G", "H"]:
            sheet.column_dimensions[col].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    period_label = f"{period_start.strftime('%Y-%m-%d') if period_start else 'inicio'}_a_{period_end.strftime('%Y-%m-%d') if period_end else 'hoy'}"
    filename = f"reportes_{period_label}.xlsx"

    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)
