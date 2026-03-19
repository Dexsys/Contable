from __future__ import annotations

import argparse
import datetime as dt
import decimal
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from app import create_app
from extensions import db
from models import Account, LedgerEntry


def normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def parse_decimal(value) -> decimal.Decimal:
    if value is None:
        return decimal.Decimal("0")
    if isinstance(value, (int, float, decimal.Decimal)):
        return decimal.Decimal(str(value))

    text = str(value).strip()
    if not text:
        return decimal.Decimal("0")

    text = text.replace("$", "").replace(" ", "")

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text == "-":
        return decimal.Decimal("0")

    return decimal.Decimal(text)


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def find_sheet(workbook, preferred: str | None, keywords: list[str], fallback_index: int):
    if preferred:
        if preferred not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja '{preferred}' en el Excel")
        return workbook[preferred]

    normalized_keywords = [normalize_text(k) for k in keywords]
    for name in workbook.sheetnames:
        normalized_name = normalize_text(name)
        if any(key in normalized_name for key in normalized_keywords):
            return workbook[name]

    return workbook[workbook.sheetnames[fallback_index]]


def first_non_empty_row(sheet):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if any(cell not in (None, "") for cell in row):
            return row_idx, row
    return None, None


def find_header_row(sheet, aliases: list[str], max_scan_rows: int = 80):
    alias_set = set(aliases)
    for row_idx in range(1, min(sheet.max_row, max_scan_rows) + 1):
        row = [sheet.cell(row_idx, col_idx).value for col_idx in range(1, sheet.max_column + 1)]
        header_map = index_headers(row)
        if any(alias in header_map for alias in alias_set):
            return row_idx, row, header_map

    row_idx, row = first_non_empty_row(sheet)
    if row is None:
        return None, None, {}
    return row_idx, row, index_headers(row)


def index_headers(header_row):
    header_map = {}
    for idx, header in enumerate(header_row):
        if header is None:
            continue
        key = normalize_text(str(header))
        if key:
            header_map[key] = idx
    return header_map


def find_column(header_map, aliases):
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]
    return None


def split_code_name(value):
    if value is None:
        return None, None
    text = str(value).strip()
    if not text:
        return None, None

    if " - " in text:
        code_part, name_part = text.split(" - ", 1)
        code = code_part.strip().replace(" ", "")
        name = name_part.strip() or None
        return code or None, name

    raw = text.replace(" ", "")
    return raw, None


def category_from_code(code: str) -> str:
    root = code.split(".")[0].strip()
    if root == "1":
        return "ingresos"
    if root == "2":
        return "egresos"
    if root == "3":
        return "banco"
    return "otros"


def parent_code(code: str) -> str | None:
    if "." not in code:
        return None
    parts = code.split(".")
    return ".".join(parts[:-1])


def upsert_account(code: str, name: str, category: str):
    existing = Account.query.filter_by(code=code).one_or_none()
    lvl = code.count(".") + 1
    parent = None
    pcode = parent_code(code)
    if pcode:
        parent = Account.query.filter_by(code=pcode).one_or_none()

    if existing:
        existing.name = name
        existing.category = category
        existing.level = lvl
        existing.parent = parent
        existing.is_postable = True
        return existing

    account = Account(
        code=code,
        name=name,
        category=category,
        level=lvl,
        parent=parent,
        is_postable=True,
    )
    db.session.add(account)
    return account


def resolve_account(raw_code: str | None, raw_name: str | None):
    if raw_code:
        by_code = Account.query.filter_by(code=raw_code).one_or_none()
        if by_code:
            return by_code

    if raw_name:
        by_name = Account.query.filter(Account.name.ilike(raw_name)).first()
        if by_name:
            return by_name

    return None


def import_plan_accounts(sheet) -> int:
    header_aliases = [
        "cuenta",
        "descripcion cuenta",
        "codigo cta",
        "codigo cuenta",
        "codigo",
    ]
    header_row_idx, header_row, header_map = find_header_row(sheet, header_aliases)
    if not header_row:
        return 0

    code_idx = find_column(
        header_map,
        ["cuenta", "codigo cta", "codigo cuenta", "codigo", "cod cuenta", "cod"],
    )
    name_idx = find_column(
        header_map,
        ["descripcion cuenta", "nombre cuenta", "descripcion", "detalle"],
    )
    composite_idx = find_column(header_map, ["codigo cta", "cuenta"])

    if code_idx is None:
        code_idx = 0
    if name_idx is None:
        name_idx = 1

    imported = 0
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        raw_code = row[code_idx] if code_idx < len(row) else None
        raw_name = row[name_idx] if name_idx < len(row) else None

        if raw_code is None and raw_name is None:
            continue

        code = str(raw_code).strip() if raw_code is not None else ""
        name = str(raw_name).strip() if raw_name is not None else ""

        if composite_idx is not None and composite_idx < len(row):
            split_code, split_name = split_code_name(row[composite_idx])
            if split_code and (not code or code == "None"):
                code = split_code
            if split_name and not name:
                name = split_name

        if not code or not name:
            continue

        code = code.replace(" ", "")
        category = category_from_code(code)
        upsert_account(code=code, name=name, category=category)
        imported += 1

        if imported % 200 == 0:
            db.session.flush()

    db.session.commit()
    return imported


def import_ledger(sheet) -> int:
    header_aliases = [
        "fecha",
        "descripcion",
        "ctactbl",
        "egresos",
        "ingresos",
        "cargos",
    ]
    header_row_idx, header_row, header_map = find_header_row(sheet, header_aliases)
    if not header_row:
        return 0

    date_idx = find_column(header_map, ["fecha", "fec", "date"])
    desc_idx = find_column(header_map, ["glosa", "descripcion", "detalle", "concepto"])
    debit_idx = find_column(header_map, ["egresos", "cargo", "cargos", "egreso", "debe", "debitos"])
    credit_idx = find_column(header_map, ["ingresos", "abono", "ingreso", "haber", "creditos"])
    cargo_idx = find_column(header_map, ["cargos", "cargo"])
    code_idx = find_column(header_map, ["ctactbl", "cuenta", "codigo cuenta", "cod cuenta", "cod"])
    account_name_idx = find_column(header_map, ["nombre cuenta", "descripcion cuenta", "cuenta nombre"])
    reference_idx = find_column(header_map, ["n doc", "folio", "referencia", "nro", "numero"])
    note_idx = find_column(header_map, ["observaciones", "observacion", "nota"])

    if date_idx is None:
        raise ValueError("No se encontró columna de fecha para movimientos históricos")

    imported = 0
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        raw_date = row[date_idx] if date_idx < len(row) else None
        entry_date = parse_date(raw_date)
        if entry_date is None:
            continue

        description = ""
        if desc_idx is not None and desc_idx < len(row) and row[desc_idx] is not None:
            description = str(row[desc_idx]).strip()

        raw_name = None
        raw_code = None
        if code_idx is not None and code_idx < len(row) and row[code_idx] is not None:
            split_code, split_name = split_code_name(row[code_idx])
            raw_code = split_code
            if split_name and (account_name_idx is None):
                raw_name = split_name

        if account_name_idx is not None and account_name_idx < len(row) and row[account_name_idx] is not None:
            raw_name = str(row[account_name_idx]).strip()

        debit_raw = row[debit_idx] if debit_idx is not None and debit_idx < len(row) else None
        credit_raw = row[credit_idx] if credit_idx is not None and credit_idx < len(row) else None
        debit = parse_decimal(debit_raw)
        credit = parse_decimal(credit_raw)

        if debit == decimal.Decimal("0") and credit == decimal.Decimal("0") and cargo_idx is not None and cargo_idx < len(row):
            cargo_value = parse_decimal(row[cargo_idx])
            if cargo_value < 0:
                debit = abs(cargo_value)
            elif cargo_value > 0:
                credit = cargo_value

        if not description:
            description = raw_name or "Movimiento sin descripción"

        if debit == decimal.Decimal("0") and credit == decimal.Decimal("0") and not description:
            continue

        reference = None
        if reference_idx is not None and reference_idx < len(row) and row[reference_idx] is not None:
            reference = str(row[reference_idx]).strip()

        note = None
        if note_idx is not None and note_idx < len(row) and row[note_idx] is not None:
            note = str(row[note_idx]).strip()

        account = resolve_account(raw_code=raw_code, raw_name=raw_name)

        item = LedgerEntry(
            entry_date=entry_date,
            bank_effective_date=entry_date,
            description=description,
            reference=reference,
            debit=debit,
            credit=credit,
            account=account,
            raw_account_code=raw_code,
            raw_account_name=raw_name,
            note=note,
            source_sheet=sheet.title,
            source_row=row_num,
            movement_type="historical_import",
        )
        db.session.add(item)
        imported += 1

        if imported % 500 == 0:
            db.session.flush()

    db.session.commit()
    return imported


def parse_args():
    parser = argparse.ArgumentParser(description="Importa plan de cuentas y datos históricos desde Excel")
    parser.add_argument("--file", required=True, help="Ruta del archivo Excel")
    parser.add_argument("--plan-sheet", help="Nombre de hoja plan de cuentas")
    parser.add_argument("--ledger-sheet", help="Nombre de hoja movimientos históricos")
    parser.add_argument("--reset", action="store_true", help="Limpia cuentas y movimientos antes de importar")
    parser.add_argument("--only-plan", action="store_true", help="Importa solo plan de cuentas")
    parser.add_argument("--only-ledger", action="store_true", help="Importa solo movimientos")
    return parser.parse_args()


def main():
    args = parse_args()
    excel_path = Path(args.file)
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {excel_path}")

    workbook = load_workbook(filename=excel_path, data_only=True)
    plan_sheet = find_sheet(workbook, args.plan_sheet, ["plan", "cuentas", "catalogo"], 0)
    ledger_sheet = find_sheet(workbook, args.ledger_sheet, ["banco", "histor", "mov", "contab", "libro"], 1 if len(workbook.sheetnames) > 1 else 0)

    app = create_app()
    with app.app_context():
        if args.reset:
            db.session.query(LedgerEntry).delete()
            db.session.query(Account).delete()
            db.session.commit()

        imported_plan = 0
        imported_ledger = 0

        if not args.only_ledger:
            imported_plan = import_plan_accounts(plan_sheet)

        if not args.only_plan:
            imported_ledger = import_ledger(ledger_sheet)

        print(f"Plan de cuentas importado: {imported_plan}")
        print(f"Movimientos históricos importados: {imported_ledger}")


if __name__ == "__main__":
    main()
